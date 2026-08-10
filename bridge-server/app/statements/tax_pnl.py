"""Parses PaytmMoney's "Tax P&L Statement" export (.xlsx), Equity sheet only
(F&O / Mutual Fund sheets out of scope — task 21). PaytmMoney has already
done the FIFO lot-matching: each row here is one already-realized lot with
its own buy/sell date+price, pre-classified into intraday/short-term/
long-term sections. We just parse and store it.

Real layout verified in task 20: three sections in one sheet, each a
section-label row (col A = label, col B = None — distinct from the summary
line above with the same label text *and* a numeric value in col B), then a
column-header row, then data rows (some rows omit the repeated "Quarter"
label — carry the last-seen value forward), then a "Total" row closing the
section."""

from datetime import datetime
from io import BytesIO

import openpyxl

SECTION_TERMS = {
    "Intraday Net Profit": "intraday",
    "Short-Term Net Profit": "short_term",
    "Long-Term Net Profit": "long_term",
}

LOT_HEADER = (
    "Quarter",
    "Scrip Name",
    "ISIN",
    "Quantity",
    "Buy Date",
    "Buy Price",
    "Buy Value",
    "Sell Date",
    "Sell Price",
    "Sell Value",
    "Net Realized P&L",
    "Brokerage",
    "Service Tax",
    "STT",
    "ETT",
    "SEBI Tax",
    "Stamp Duty",
    "Total Charges & Tax",
)


class TaxPnlParseError(ValueError):
    pass


def _parse_date(raw) -> "datetime.date":
    if isinstance(raw, datetime):
        return raw.date()
    return datetime.strptime(str(raw).strip(), "%d-%b-%Y").date()


def _is_section_label(row: tuple) -> str | None:
    if row and row[0] in SECTION_TERMS and (len(row) < 2 or row[1] is None):
        return SECTION_TERMS[row[0]]
    return None


def parse_tax_pnl(file_bytes: bytes) -> tuple[str, list[dict]]:
    """Returns (financial_year, lots)."""
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise TaxPnlParseError(f"Not a readable .xlsx file: {e}") from e
    if "Equity" not in wb.sheetnames:
        raise TaxPnlParseError("No 'Equity' sheet found — unexpected file format")
    ws = wb["Equity"]
    rows = list(ws.iter_rows(values_only=True))

    financial_year = next(
        (str(row[1]) for row in rows if row and row[0] == "Period" and row[1]),
        None,
    )
    if not financial_year:
        raise TaxPnlParseError("Could not find the 'Period' row — unexpected file format")

    lots: list[dict] = []
    i = 0
    while i < len(rows):
        term = _is_section_label(rows[i])
        if term is None:
            i += 1
            continue

        # Next non-blank row after a section label is the column-header row.
        j = i + 1
        while j < len(rows) and (not rows[j] or rows[j][0] is None):
            j += 1
        if j >= len(rows) or rows[j][: len(LOT_HEADER)] != LOT_HEADER:
            raise TaxPnlParseError(f"Expected lot header after {term!r} section label, row {j + 1}")

        current_quarter = None
        k = j + 1
        while k < len(rows):
            row = rows[k]
            # A blank Quarter (col 0) with a populated Scrip Name (col 1) is a
            # valid continuation row carrying the last quarter forward — only
            # a row blank in *both* is a genuine separator. Blank rows only
            # ever separate quarter groups within a section (verified against
            # real data); "Total" is the sole terminator, so skip blanks
            # rather than treating them as end-of-section.
            if not row or (row[0] is None and (len(row) < 2 or row[1] is None)):
                k += 1
                continue
            if row[0] == "Total":
                k += 1
                break
            if row[0] in SECTION_TERMS and (len(row) < 2 or row[1] is None):
                break  # next section label reached without a "Total" row — stop, don't consume it
            (
                quarter,
                scrip_name,
                isin,
                quantity,
                buy_date,
                buy_price,
                buy_value,
                sell_date,
                sell_price,
                sell_value,
                net_pnl,
                brokerage,
                service_tax,
                stt,
                ett,
                sebi_tax,
                stamp_duty,
                total_charges,
            ) = row[: len(LOT_HEADER)]
            if quarter:
                current_quarter = str(quarter)
            lots.append(
                {
                    "term": term,
                    "quarter": current_quarter,
                    "scrip_name": str(scrip_name),
                    "isin": str(isin),
                    "quantity": float(quantity),
                    "buy_date": _parse_date(buy_date),
                    "buy_price": float(buy_price),
                    "buy_value": float(buy_value),
                    "sell_date": _parse_date(sell_date),
                    "sell_price": float(sell_price),
                    "sell_value": float(sell_value),
                    "net_realized_pnl": float(net_pnl),
                    "brokerage": float(brokerage or 0),
                    "service_tax": float(service_tax or 0),
                    "stt": float(stt or 0),
                    "ett": float(ett or 0),
                    "sebi_tax": float(sebi_tax or 0),
                    "stamp_duty": float(stamp_duty or 0),
                    "total_charges": float(total_charges or 0),
                }
            )
            k += 1
        i = k

    return financial_year, lots
